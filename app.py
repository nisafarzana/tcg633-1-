# =============================================================================
# Digital Pavement Condition Evaluation and Maintenance Decision Tool
# TCG633 Bridge & Road Maintenance — Individual Project
# Universiti Teknologi MARA, Cawangan Sarawak
#
# Computation logic (PCI weighting/severity factors, classification bands,
# IRI bands, and the PCI/IRI hybrid rule) is taken directly from the
# lecturer-provided files:
#   - TCG633_PCI_IRI_Model.xlsx      (Lookup, PCI_Compute, IRI_Compute sheets)
#   - TCG633_PCI_IRI_Pro_v2.xlsx     (Lookup, Settings_Summary "Hybrid" logic)
# Any place where the spreadsheet logic does not cover a case (e.g. a defect
# type or severity not in the Lookup table) is explicitly flagged as an
# ASSUMPTION in-app (see "Methodology & Assumptions" page) rather than
# silently guessed.
#
# This file uses Streamlit's native multi-page navigation (st.navigation /
# st.Page) instead of st.tabs. The practical benefit: only the page the user
# is currently viewing actually executes and renders — switching pages does
# not re-run the chart/table-building code of every other page, which is
# what made the previous tabs-based layout feel sluggish on larger datasets.
# =============================================================================

import io
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st

# Plotly is an optional, "nice-to-have" charting dependency. If the deployment
# environment fails to install it for any reason (e.g. requirements.txt not
# picked up yet on Streamlit Cloud), the app must NOT crash — it should fall
# back to Streamlit's built-in charts instead.
try:
    import plotly.express as px
    PLOTLY_OK = True
except ModuleNotFoundError:
    PLOTLY_OK = False

# openpyxl powers BOTH reading uploaded .xlsx/.xls files AND writing the Excel
# download. If it fails to import (e.g. a broken Cloud build environment),
# the app must still work for CSV upload/download — it must never crash.
try:
    import openpyxl  # noqa: F401
    OPENPYXL_OK = True
except ModuleNotFoundError:
    OPENPYXL_OK = False

# -----------------------------------------------------------------------------
# PAGE CONFIG  (must be the very first Streamlit command)
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="Pavement Condition Evaluation & Maintenance Decision Tool",
    page_icon="🛣️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -----------------------------------------------------------------------------
# LOOK-UP CONSTANTS  (copied from the lecturer's Lookup sheet)
# -----------------------------------------------------------------------------

# PCI defect weighting factors — Lookup!A:B
DEFECT_WEIGHTS = {
    "Longitudinal Crack": 1.0,
    "Alligator (Fatigue) Crack": 1.6,
    "Potholes": 2.2,
    "Raveling": 1.2,
    "Depression/Sag": 1.4,
    "Patching (Failed)": 1.8,
    "Bleeding/Flushing": 1.0,
    "Rut/Rutting": 1.6,
}
DEFAULT_WEIGHT = 1.0  # ASSUMPTION: used only if an unrecognised defect type is supplied

# PCI severity factors — Lookup!D:E
SEVERITY_FACTORS = {"Low": 0.6, "Medium": 1.0, "High": 1.4}
DEFAULT_SEVERITY_FACTOR = 1.0  # ASSUMPTION: "Medium" weighting used if severity is unrecognised

# PCI condition bands — Lookup!G:J  (rank 1 = best ... 4 = worst)
PCI_RANK_LABEL = {1: "Very Good", 2: "Good / Satisfactory", 3: "Fair", 4: "Poor"}
PCI_RANK_RECO = {
    1: "Routine maintenance (cleaning, grass cutting, minor touch-ups)",
    2: "Preventive maintenance (crack sealing, local patching)",
    3: "Surface treatment / Overlay (localized)",
    4: "Major rehabilitation / Reconstruction assessment",
}

# IRI condition bands — Lookup!L:O (rank 1 = best ... 4 = worst)
IRI_RANK_LABEL = {1: "Very Good (Smooth)", 2: "Good", 3: "Fair", 4: "Poor (Rough)"}
IRI_RANK_RECO = {
    1: "Routine maintenance",
    2: "Preventive maintenance (localized patching/leveling)",
    3: "Surface treatment / thin overlay",
    4: "Structural overlay / rehabilitation",
}

CONDITION_COLORS = {
    "Very Good": "#2E7D32",
    "Very Good (Smooth)": "#2E7D32",
    "Good": "#1976D2",
    "Good / Satisfactory": "#1976D2",
    "Fair": "#F9A825",
    "Poor": "#C62828",
    "Poor (Rough)": "#C62828",
}
RANK_COLOR = {1: "#2E7D32", 2: "#1976D2", 3: "#F9A825", 4: "#C62828"}

# Supplementary, defect-level treatment guide.
# NOTE / ASSUMPTION: the lecturer's spreadsheet only issues a maintenance
# recommendation at the SECTION level (based on PCI/IRI classification).
# It does not provide a per-defect-type x severity action table. The table
# below is added as general pavement-maintenance practice guidance (common,
# non-proprietary treatments) so that the tool can also recommend an action
# "for each defect" as requested. It is clearly separated from the official
# section-level recommendation in every table/report this app produces.
DEFECT_TREATMENT_GUIDE = {
    "Longitudinal Crack": {
        "Low": "Monitor; seal during routine maintenance",
        "Medium": "Crack sealing",
        "High": "Crack sealing + localized patching",
    },
    "Alligator (Fatigue) Crack": {
        "Low": "Monitor / crack sealing",
        "Medium": "Partial-depth patching",
        "High": "Full-depth patching or overlay (structural distress)",
    },
    "Potholes": {
        "Low": "Patch at next routine maintenance round",
        "Medium": "Patch promptly",
        "High": "Immediate patching — safety hazard",
    },
    "Raveling": {
        "Low": "Monitor / fog seal",
        "Medium": "Surface (chip/slurry) seal",
        "High": "Thin overlay",
    },
    "Depression/Sag": {
        "Low": "Monitor drainage and surface",
        "Medium": "Localized levelling / patching",
        "High": "Investigate sub-base; structural repair",
    },
    "Patching (Failed)": {
        "Low": "Reseal patch edges",
        "Medium": "Remove and re-patch",
        "High": "Full-depth repair of patch area",
    },
    "Bleeding/Flushing": {
        "Low": "Apply sand/blotter material",
        "Medium": "Surface treatment",
        "High": "Overlay (loss of skid resistance)",
    },
    "Rut/Rutting": {
        "Low": "Monitor",
        "Medium": "Milling and overlay",
        "High": "Structural overlay / reconstruction",
    },
}
DEFAULT_TREATMENT = "Inspect on-site and patch/repair as needed (defect type not in guide)"

CANONICAL_COLS = ["Section", "Defect Type", "Severity", "Area Percentage (%)", "IRI"]

# Map of common header variants -> canonical column name
COLUMN_ALIASES = {
    "section": "Section", "section id": "Section", "sectionid": "Section",
    "section_id": "Section", "road section": "Section",
    "defect type": "Defect Type", "defect": "Defect Type", "defecttype": "Defect Type",
    "distress type": "Defect Type", "distress": "Defect Type",
    "severity": "Severity", "severity level": "Severity",
    "area percentage (%)": "Area Percentage (%)", "area percentage": "Area Percentage (%)",
    "area (%)": "Area Percentage (%)", "area affected (%)": "Area Percentage (%)",
    "area affected": "Area Percentage (%)", "area%": "Area Percentage (%)",
    "area_pct": "Area Percentage (%)", "area": "Area Percentage (%)",
    "iri": "IRI", "iri (m/km)": "IRI", "avg iri": "IRI", "average iri": "IRI",
    "iri(m/km)": "IRI", "roughness": "IRI",
}


# -----------------------------------------------------------------------------
# HELPER FUNCTIONS — DATA / PARSING / COMPUTATION
# (Unchanged from the previous version; this is the tested, verified engine.)
# -----------------------------------------------------------------------------
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns to the canonical set using a case/space-insensitive match."""
    rename_map = {}
    for col in df.columns:
        key = str(col).strip().lower()
        if key in COLUMN_ALIASES:
            rename_map[col] = COLUMN_ALIASES[key]
        elif str(col).strip() in CANONICAL_COLS:
            rename_map[col] = str(col).strip()
    return df.rename(columns=rename_map)


def pci_rank(pci: float) -> int:
    if pci >= 85:
        return 1
    if pci >= 70:
        return 2
    if pci >= 55:
        return 3
    return 4


def iri_rank(iri: float) -> int:
    if iri < 2:
        return 1
    if iri < 3:
        return 2
    if iri < 4:
        return 3
    return 4


def lookup_weight(defect_type, flags):
    key = str(defect_type).strip()
    for k, v in DEFECT_WEIGHTS.items():
        if k.lower() == key.lower():
            return v
    flags.append(f"Unrecognised defect type '{defect_type}' — default weighting {DEFAULT_WEIGHT} applied")
    return DEFAULT_WEIGHT


def lookup_severity(severity, flags):
    key = str(severity).strip()
    for k, v in SEVERITY_FACTORS.items():
        if k.lower() == key.lower():
            return v
    flags.append(f"Unrecognised severity '{severity}' — default factor {DEFAULT_SEVERITY_FACTOR} (Medium) applied")
    return DEFAULT_SEVERITY_FACTOR


def lookup_treatment(defect_type, severity):
    for k, v in DEFECT_TREATMENT_GUIDE.items():
        if k.lower() == str(defect_type).strip().lower():
            for sk, sv in v.items():
                if sk.lower() == str(severity).strip().lower():
                    return sv
            return DEFAULT_TREATMENT
    return DEFAULT_TREATMENT


@st.cache_data
def generate_sample_data() -> pd.DataFrame:
    """Built-in dataset extracted directly from the student's own completed
    workbook (`TCG633_PCI_IRI_Pro_v2_completed.xlsx` — PCI_Input + IRI_Input
    sheets, Sections 1-10). IRI is the average of that section's 10 roughness
    segments from IRI_Input. This lets "Load Sample Data" work immediately
    even on a server where Excel upload is broken (no openpyxl needed),
    while still showing the assignment's actual dataset rather than a
    generic placeholder."""
    rows = [
        # Section, Defect Type, Severity, Area %, IRI (m/km, avg of 10 segments)
        ("S1", "Longitudinal Crack", "Low", 5, 1.60),
        ("S2", "Potholes", "Medium", 3, 1.91),
        ("S3", "Raveling", "Low", 10, 2.21),
        ("S4", "Alligator (Fatigue) Crack", "Medium", 6, 2.51),
        ("S5", "Depression/Sag", "Low", 4, 2.01),
        ("S6", "Patching (Failed)", "High", 2, 2.81),
        ("S7", "Alligator (Fatigue) Crack", "High", 10, 3.31),
        ("S8", "Potholes", "High", 11, 3.81),
        ("S9", "Patching (Failed)", "High", 20, 4.51),
        ("S10", "Rut/Rutting", "High", 30, 5.51),
    ]
    return pd.DataFrame(rows, columns=CANONICAL_COLS)


def _is_lecturer_template(wb) -> bool:
    """Return True if this workbook is the lecturer's TCG633 template format
    (has both PCI_Input and IRI_Input sheets)."""
    return "PCI_Input" in wb.sheetnames and "IRI_Input" in wb.sheetnames


def _parse_lecturer_template(wb) -> pd.DataFrame:
    """Parse the lecturer's TCG633_PCI_IRI_Pro workbook format.

    PCI_Input  — header on row 6, data from row 7.
                 Columns: Section ID | Defect Type | Severity | Area Affected (%) | Notes
                 Only rows where *all* of Defect Type, Severity AND Area Affected are
                 non-null are considered actual defect records.

    IRI_Input  — header on row 6, data from row 7.
                 Columns: Section ID | Segment ID | Start | End | IRI (m/km) | Notes
                 IRI values are averaged per Section to produce one IRI value per section.

    Both sheets are merged on Section ID into the flat 5-column format the
    computation engine expects: Section, Defect Type, Severity, Area Percentage (%), IRI.
    """
    # ---------- PCI_Input ----------
    ws_pci = wb["PCI_Input"]
    pci_rows = []
    for row in ws_pci.iter_rows(min_row=7, values_only=True):
        sec_id, defect, severity, area, *_ = row + (None,) * 5
        if sec_id is None:
            continue
        if defect is None or severity is None or area is None:
            continue  # skip placeholder / empty rows
        pci_rows.append({
            "Section": f"S{int(sec_id)}" if isinstance(sec_id, (int, float)) else str(sec_id).strip(),
            "Defect Type": str(defect).strip(),
            "Severity": str(severity).strip(),
            "Area Percentage (%)": float(area),
        })
    pci_df = pd.DataFrame(pci_rows) if pci_rows else pd.DataFrame(
        columns=["Section", "Defect Type", "Severity", "Area Percentage (%)"]
    )

    # ---------- IRI_Input ----------
    ws_iri = wb["IRI_Input"]
    iri_rows = []
    for row in ws_iri.iter_rows(min_row=7, values_only=True):
        sec_id = row[0]
        iri_val = row[4] if len(row) > 4 else None
        if sec_id is None or iri_val is None:
            continue
        try:
            iri_val = float(iri_val)
        except (TypeError, ValueError):
            continue
        sec_label = f"S{int(sec_id)}" if isinstance(sec_id, (int, float)) else str(sec_id).strip()
        iri_rows.append({"Section": sec_label, "IRI": iri_val})
    iri_df = pd.DataFrame(iri_rows) if iri_rows else pd.DataFrame(columns=["Section", "IRI"])
    avg_iri = iri_df.groupby("Section")["IRI"].mean().reset_index() if not iri_df.empty else pd.DataFrame(columns=["Section", "IRI"])

    # ---------- Merge ----------
    all_sections = sorted(
        set(pci_df["Section"].tolist()) | set(avg_iri["Section"].tolist()),
        key=lambda x: (len(x), x)
    )
    merged_rows = []
    for sec in all_sections:
        sec_pci = pci_df[pci_df["Section"] == sec]
        iri_val = avg_iri.loc[avg_iri["Section"] == sec, "IRI"].values
        iri_val = float(iri_val[0]) if len(iri_val) > 0 else np.nan
        if not sec_pci.empty:
            for _, r in sec_pci.iterrows():
                merged_rows.append({
                    "Section": sec,
                    "Defect Type": r["Defect Type"],
                    "Severity": r["Severity"],
                    "Area Percentage (%)": r["Area Percentage (%)"],
                    "IRI": iri_val,
                })
        else:
            # IRI-only section
            merged_rows.append({
                "Section": sec,
                "Defect Type": np.nan,
                "Severity": np.nan,
                "Area Percentage (%)": np.nan,
                "IRI": iri_val,
            })
    return pd.DataFrame(merged_rows, columns=CANONICAL_COLS)


def read_uploaded_file(uploaded_file) -> pd.DataFrame:
    """Smart parser that handles:
    1. CSV files (any flat layout) — no openpyxl needed.
    2. The lecturer's TCG633 multi-sheet Excel template (PCI_Input + IRI_Input sheets).
    3. Any generic flat Excel file with the 5-column layout.
    """
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
        return normalize_columns(df)

    # Excel path — needs openpyxl
    if not OPENPYXL_OK:
        raise RuntimeError(
            "This server's Python environment is missing the 'openpyxl' package, "
            "so Excel (.xlsx/.xls) files can't be read right now. "
            "Workaround: open this file in Excel/Google Sheets and re-save/export it "
            "as .csv, then upload the .csv instead — CSV upload does not need openpyxl."
        )

    import openpyxl as _openpyxl
    # Read bytes once so we can open with openpyxl AND pandas without seeking issues
    raw_bytes = uploaded_file.read()
    wb = _openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)

    if _is_lecturer_template(wb):
        return _parse_lecturer_template(wb)

    # Generic flat Excel — try first sheet
    df = pd.read_excel(io.BytesIO(raw_bytes), sheet_name=0)
    return normalize_columns(df)


@st.cache_data(show_spinner=False)
def compute_results(df: pd.DataFrame):
    """Core engine: replicates PCI_Compute / IRI_Compute / Settings_Summary
    logic from the lecturer's Excel model. Returns (detail_df, summary_df, flags).
    Cached on the exact input dataframe so re-visiting a page after switching
    elsewhere doesn't recompute from scratch."""
    flags = []
    df = df.copy()

    for c in CANONICAL_COLS:
        if c not in df.columns:
            df[c] = np.nan

    df["Section"] = df["Section"].astype(str).str.strip()
    df = df[df["Section"].notna() & (df["Section"] != "") & (df["Section"].str.lower() != "nan")]

    df["Area Percentage (%)"] = pd.to_numeric(df["Area Percentage (%)"], errors="coerce")
    df["IRI"] = pd.to_numeric(df["IRI"], errors="coerce")

    defect_mask = (
        df["Defect Type"].notna()
        & df["Severity"].notna()
        & df["Area Percentage (%)"].notna()
    )
    detail = df[defect_mask].copy()

    if not detail.empty:
        detail["Weighting Factor"] = detail["Defect Type"].apply(lambda d: lookup_weight(d, flags))
        detail["Severity Factor"] = detail["Severity"].apply(lambda s: lookup_severity(s, flags))
        detail["Deduct Value"] = (
            detail["Area Percentage (%)"] * detail["Severity Factor"] * detail["Weighting Factor"]
        )
        detail["Suggested Defect Treatment"] = detail.apply(
            lambda r: lookup_treatment(r["Defect Type"], r["Severity"]), axis=1
        )
    else:
        for c in ["Weighting Factor", "Severity Factor", "Deduct Value", "Suggested Defect Treatment"]:
            detail[c] = np.nan

    all_sections = pd.Index(sorted(df["Section"].unique(), key=lambda x: (len(x), x)))

    sum_deduct = detail.groupby("Section")["Deduct Value"].sum() if not detail.empty else pd.Series(dtype=float)
    defect_count = detail.groupby("Section").size() if not detail.empty else pd.Series(dtype=int)
    avg_iri = df[df["IRI"].notna()].groupby("Section")["IRI"].mean()

    rows = []
    for sec in all_sections:
        has_pci = sec in sum_deduct.index
        has_iri = sec in avg_iri.index

        pci_val = max(0.0, 100.0 - min(100.0, sum_deduct.get(sec, 0.0))) if has_pci else np.nan
        iri_val = float(avg_iri.get(sec)) if has_iri else np.nan

        r_pci = pci_rank(pci_val) if has_pci else None
        r_iri = iri_rank(iri_val) if has_iri else None

        if has_pci and has_iri:
            combined_rank = max(r_pci, r_iri)
            combined_label = PCI_RANK_LABEL[combined_rank]
            combined_reco = PCI_RANK_RECO[combined_rank]
            basis = "Hybrid (PCI & IRI — worse of the two governs)"
        elif has_pci:
            combined_rank = r_pci
            combined_label = PCI_RANK_LABEL[combined_rank]
            combined_reco = PCI_RANK_RECO[combined_rank]
            basis = "PCI only (no IRI data)"
        elif has_iri:
            combined_rank = r_iri
            combined_label = IRI_RANK_LABEL[combined_rank]
            combined_reco = IRI_RANK_RECO[combined_rank]
            basis = "IRI only (no defect data)"
        else:
            combined_rank, combined_label, combined_reco, basis = None, "No Data", "—", "No data"

        rows.append({
            "Section": sec,
            "No. of Defects Recorded": int(defect_count.get(sec, 0)),
            "Sum Deduct Value": round(sum_deduct.get(sec, np.nan), 2) if has_pci else np.nan,
            "PCI": round(pci_val, 1) if has_pci else np.nan,
            "PCI Condition": PCI_RANK_LABEL[r_pci] if has_pci else "—",
            "Avg IRI (m/km)": round(iri_val, 2) if has_iri else np.nan,
            "IRI Condition": IRI_RANK_LABEL[r_iri] if has_iri else "—",
            "Combined Condition Rating": combined_label,
            "_rank": combined_rank,
            "Maintenance Recommendation": combined_reco,
            "Basis": basis,
        })

    summary = pd.DataFrame(rows)
    return detail, summary, flags


def df_to_excel_bytes(sheets: dict) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, d in sheets.items():
            d.to_excel(writer, sheet_name=name[:31], index=False)
    return buf.getvalue()


def condition_badge(label: str) -> str:
    color = CONDITION_COLORS.get(label, "#777")
    return f'<span class="badge" style="background:{color}">{label}</span>'


# Colour maps used by style_dataframe
SEVERITY_COLORS = {
    "Low":    {"bg": "#E8F5E9", "fg": "#1B5E20"},   # soft green
    "Medium": {"bg": "#FFF8E1", "fg": "#F57F17"},   # soft amber
    "High":   {"bg": "#FFEBEE", "fg": "#B71C1C"},   # soft red
}

CONDITION_BG = {
    "Very Good":          {"bg": "#E8F5E9", "fg": "#1B5E20"},
    "Very Good (Smooth)": {"bg": "#E8F5E9", "fg": "#1B5E20"},
    "Good":               {"bg": "#E3F2FD", "fg": "#0D47A1"},
    "Good / Satisfactory":{"bg": "#E3F2FD", "fg": "#0D47A1"},
    "Fair":               {"bg": "#FFF8E1", "fg": "#E65100"},
    "Poor":               {"bg": "#FFEBEE", "fg": "#B71C1C"},
    "Poor (Rough)":       {"bg": "#FFEBEE", "fg": "#B71C1C"},
}


def _cell_color(val, mapping: dict) -> str:
    """Return a pandas Styler CSS string for a cell value looked up in mapping."""
    entry = mapping.get(str(val).strip(), {})
    if not entry:
        return ""
    return f"background-color: {entry['bg']}; color: {entry['fg']}; font-weight: 600;"


def style_dataframe(df: pd.DataFrame) -> "pd.io.formats.style.Styler":
    """Apply colour coding to condition label and severity columns only.
    Numeric columns (PCI, IRI values, Hybrid Index) are left uncoloured
    so the table stays clean and easy to read.
    """
    styler = df.style

    # pandas ≥2.1 renamed applymap → map; support both
    _applymap = getattr(styler, "map", None) or getattr(styler, "applymap")

    # --- Severity labels ---
    if "Severity" in df.columns:
        styler = _applymap(
            lambda v: _cell_color(v, SEVERITY_COLORS), subset=["Severity"]
        )

    # --- Condition label columns (any column with "Condition" in the name) ---
    cond_cols = [c for c in df.columns if "Condition" in c]
    for col in cond_cols:
        styler = _applymap(
            lambda v: _cell_color(v, CONDITION_BG), subset=[col]
        )

    return styler


# -----------------------------------------------------------------------------
# BONUS FEATURE 1 — HYBRID INDEX (numeric PCI + IRI blend, 0-100 scale)
# -----------------------------------------------------------------------------
# NOTE: this is an ADDITIONAL, OPTIONAL bonus feature on top of the official
# section-level "Combined Condition Rating" computed above (which follows the
# lecturer's "worse of PCI/IRI governs" rule and is what the rest of the app
# uses by default). The Hybrid Index instead produces one continuous 0-100
# NUMBER per section by rescaling IRI onto the same 0-100 scale PCI already
# uses (anchored at the identical classification boundaries: 85/70/55), then
# blending the two with a user-adjustable weight. This satisfies the
# "combine PCI & IRI into a hybrid index" bonus request from the project
# brief, while keeping the official rating untouched and clearly labelled
# elsewhere in the app.
IRI_SCORE_ANCHORS = [  # (IRI m/km, equivalent 0-100 score) — monotonic, decreasing score
    (0.0, 100.0),
    (2.0, 85.0),
    (3.0, 70.0),
    (4.0, 55.0),
    (6.0, 0.0),
]


def iri_to_score(iri: float) -> float:
    """Piecewise-linear rescale of IRI (m/km, lower = better) onto a 0-100
    score using the SAME boundary values as the PCI classification bands
    (85 / 70 / 55), so a weighted blend with PCI is meaningful. Values above
    6 m/km clamp to 0; values below 0 clamp to 100 (shouldn't occur)."""
    if pd.isna(iri):
        return np.nan
    if iri <= IRI_SCORE_ANCHORS[0][0]:
        return IRI_SCORE_ANCHORS[0][1]
    if iri >= IRI_SCORE_ANCHORS[-1][0]:
        return IRI_SCORE_ANCHORS[-1][1]
    for (x0, y0), (x1, y1) in zip(IRI_SCORE_ANCHORS, IRI_SCORE_ANCHORS[1:]):
        if x0 <= iri <= x1:
            frac = (iri - x0) / (x1 - x0)
            return y0 + frac * (y1 - y0)
    return np.nan


def compute_hybrid_index(summary_df: pd.DataFrame, w_pci: float) -> pd.DataFrame:
    """Add 'IRI Score (0-100)', 'Hybrid Index', and 'Hybrid Condition' columns.
    w_pci is the weight given to PCI (0.0-1.0); IRI gets (1 - w_pci)."""
    out = summary_df.copy()
    out["IRI Score (0-100)"] = out["Avg IRI (m/km)"].apply(iri_to_score)

    def blend(row):
        has_pci = pd.notna(row["PCI"])
        has_iri = pd.notna(row["IRI Score (0-100)"])
        if has_pci and has_iri:
            return w_pci * row["PCI"] + (1 - w_pci) * row["IRI Score (0-100)"]
        if has_pci:
            return row["PCI"]
        if has_iri:
            return row["IRI Score (0-100)"]
        return np.nan

    out["Hybrid Index"] = out.apply(blend, axis=1).round(1)
    out["Hybrid Condition"] = out["Hybrid Index"].apply(
        lambda v: PCI_RANK_LABEL[pci_rank(v)] if pd.notna(v) else "—"
    )
    return out


# -----------------------------------------------------------------------------
# BONUS FEATURE 2 — AUTOMATED REPORT GENERATION (self-contained HTML)
# -----------------------------------------------------------------------------
# NOTE: deliberately built with ZERO new pip dependencies (no reportlab/fpdf/
# weasyprint) given how much trouble missing packages have already caused on
# Streamlit Cloud in this project. The report is a single self-contained
# .html file (tables + small hand-built inline SVG bar charts) that opens in
# any browser and can be turned into a PDF via the browser's native
# "Print > Save as PDF" — no server-side PDF engine required.
def _svg_bar_chart(labels, values, colors, title, width=640, height=260, value_fmt="{:.1f}"):
    """Tiny, dependency-free inline SVG bar chart generator for the report."""
    n = len(labels)
    if n == 0:
        return f"<p><em>No data for {title}.</em></p>"
    pad_l, pad_r, pad_t, pad_b = 40, 20, 30, 50
    chart_w = width - pad_l - pad_r
    chart_h = height - pad_t - pad_b
    max_val = max([v for v in values if pd.notna(v)] or [1]) * 1.15 or 1
    bar_w = chart_w / n * 0.6
    gap = chart_w / n

    bars = []
    for i, (lab, val, col) in enumerate(zip(labels, values, colors)):
        if pd.isna(val):
            continue
        bh = (val / max_val) * chart_h
        x = pad_l + i * gap + (gap - bar_w) / 2
        y = pad_t + (chart_h - bh)
        bars.append(
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w:.1f}" height="{bh:.1f}" '
            f'fill="{col}" rx="3" />'
            f'<text x="{x + bar_w/2:.1f}" y="{y - 5:.1f}" font-size="11" '
            f'text-anchor="middle" fill="#333">{value_fmt.format(val)}</text>'
            f'<text x="{x + bar_w/2:.1f}" y="{pad_t + chart_h + 16:.1f}" font-size="10" '
            f'text-anchor="middle" fill="#555">{lab}</text>'
        )
    svg = (
        f'<svg viewBox="0 0 {width} {height}" width="100%" style="max-width:{width}px">'
        f'<line x1="{pad_l}" y1="{pad_t + chart_h}" x2="{pad_l + chart_w}" y2="{pad_t + chart_h}" '
        f'stroke="#ccc" stroke-width="1"/>'
        f'<text x="{pad_l}" y="18" font-size="13" font-weight="600" fill="#222">{title}</text>'
        + "".join(bars) +
        "</svg>"
    )
    return svg


def build_html_report(summary_df, detail_df, data_source, calc_flags, hybrid_df=None) -> str:
    ts_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    display_summary = summary_df.drop(columns=["_rank"], errors="ignore")

    avg_pci = summary_df["PCI"].mean()
    avg_iri = summary_df["Avg IRI (m/km)"].mean()
    poor_n = int(summary_df["Combined Condition Rating"].isin(["Poor", "Poor (Rough)"]).sum())
    total_n = len(summary_df)

    pci_chart = _svg_bar_chart(
        list(summary_df["Section"]), list(summary_df["PCI"]),
        [CONDITION_COLORS.get(c, "#999") for c in summary_df["PCI Condition"]],
        "PCI by Section",
    )
    iri_chart = _svg_bar_chart(
        list(summary_df["Section"]), list(summary_df["Avg IRI (m/km)"]),
        [CONDITION_COLORS.get(c, "#999") for c in summary_df["IRI Condition"]],
        "Average IRI by Section (m/km)", value_fmt="{:.2f}",
    )

    hybrid_section = ""
    if hybrid_df is not None and not hybrid_df.empty:
        hybrid_chart = _svg_bar_chart(
            list(hybrid_df["Section"]), list(hybrid_df["Hybrid Index"]),
            [CONDITION_COLORS.get(c, "#999") for c in hybrid_df["Hybrid Condition"]],
            "Hybrid Index by Section",
        )
        hybrid_table = hybrid_df[["Section", "PCI", "IRI Score (0-100)", "Hybrid Index", "Hybrid Condition"]].to_html(
            index=False, classes="datatable", border=0
        )
        hybrid_section = f"""
        <h2>5. Hybrid Index (Bonus)</h2>
        <p>An additional numeric PCI+IRI blend on a 0-100 scale (see Methodology page for the formula).
        This supplements — and does not replace — the official Combined Condition Rating above.</p>
        {hybrid_chart}
        {hybrid_table}
        """

    flags_html = ""
    if calc_flags:
        flags_html = "<h2>6. Data-Quality Flags</h2><ul>" + "".join(
            f"<li>{f}</li>" for f in sorted(set(calc_flags))
        ) + "</ul>"

    detail_table_html = ""
    if detail_df is not None and not detail_df.empty:
        detail_cols = ["Section", "Defect Type", "Severity", "Area Percentage (%)",
                        "Weighting Factor", "Severity Factor", "Deduct Value",
                        "Suggested Defect Treatment"]
        detail_table_html = detail_df[detail_cols].to_html(index=False, classes="datatable", border=0)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Pavement Condition Evaluation Report</title>
<style>
  body {{ font-family: -apple-system, Segoe UI, Roboto, Arial, sans-serif; color: #222;
          max-width: 900px; margin: 30px auto; padding: 0 20px; line-height: 1.5; }}
  h1 {{ font-size: 26px; border-bottom: 3px solid #1976D2; padding-bottom: 10px; }}
  h2 {{ font-size: 18px; color: #1976D2; margin-top: 32px; }}
  .meta {{ color: #666; font-size: 13px; margin-bottom: 20px; }}
  .kpi-row {{ display: flex; gap: 14px; flex-wrap: wrap; margin: 18px 0; }}
  .kpi {{ background: #f5f7fa; border: 1px solid #e0e0e0; border-radius: 8px;
          padding: 12px 18px; min-width: 140px; }}
  .kpi .label {{ font-size: 12px; color: #666; }}
  .kpi .value {{ font-size: 22px; font-weight: 700; color: #1976D2; }}
  table.datatable {{ border-collapse: collapse; width: 100%; font-size: 13px; margin-top: 10px; }}
  table.datatable th {{ background: #1976D2; color: white; padding: 7px 10px; text-align: left; }}
  table.datatable td {{ padding: 6px 10px; border-bottom: 1px solid #eee; }}
  table.datatable tr:nth-child(even) {{ background: #f8fafc; }}
  .flagbox {{ background: #fff8e1; border: 1px solid #ffe082; border-radius: 6px; padding: 10px 14px; }}
  .footer {{ margin-top: 40px; font-size: 12px; color: #999; border-top: 1px solid #eee; padding-top: 12px; }}
  @media print {{ body {{ margin: 0; }} }}
</style>
</head>
<body>

<h1>🛣️ Digital Pavement Condition Evaluation Report</h1>
<div class="meta">
  TCG633 Bridge &amp; Road Maintenance · UiTM Cawangan Sarawak<br>
  Generated: {ts_str} &nbsp;|&nbsp; Data source: {data_source or "—"}
</div>

<h2>1. Summary Dashboard</h2>
<div class="kpi-row">
  <div class="kpi"><div class="label">Total Sections</div><div class="value">{total_n}</div></div>
  <div class="kpi"><div class="label">Average PCI</div><div class="value">{avg_pci:.1f}</div></div>
  <div class="kpi"><div class="label">Average IRI (m/km)</div><div class="value">{avg_iri:.2f}</div></div>
  <div class="kpi"><div class="label">Poor Sections</div><div class="value">{poor_n}</div></div>
</div>

<h2>2. Charts</h2>
{pci_chart}
{iri_chart}

<h2>3. Section-Level Summary</h2>
{display_summary.to_html(index=False, classes="datatable", border=0)}

<h2>4. Defect-Level Detail</h2>
{detail_table_html or "<p><em>No defect-level data in this dataset.</em></p>"}

{hybrid_section}

{flags_html}

<h2>7. Methodology (brief)</h2>
<p>PCI uses a deduct-value method: <code>Deduct = Area% × Severity Factor × Weighting Factor</code>,
summed per section, then <code>PCI = MAX(0, 100 − MIN(100, ΣDeduct))</code>. IRI is the
average of all roughness readings per section. The Combined Condition Rating takes the
worse of the PCI and IRI classification bands per the lecturer's hybrid rule. Full detail
is in the app's "Methodology &amp; Assumptions" page.</p>

<div class="footer">
  Auto-generated by the TCG633 Digital Pavement Condition Evaluation and Maintenance
  Decision Tool (Streamlit). Tip: use your browser's Print → Save as PDF to get a PDF copy
  of this report.
</div>

</body>
</html>"""
    return html


# -----------------------------------------------------------------------------
# LIGHT CUSTOM STYLING
# -----------------------------------------------------------------------------
st.markdown(
    """
    <style>
    .main > div {padding-top: 1.0rem;}
    .kpi-box {
        background: #ffffff; border: 1px solid #e6e6e6; border-radius: 10px;
        padding: 14px 18px; box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    }
    .small-note {color:#666; font-size:0.85rem;}
    .badge {
        display:inline-block; padding:3px 10px; border-radius:14px;
        color:white; font-size:0.78rem; font-weight:600;
    }
    [data-testid="stMetric"] {
        background: #ffffff; border: 1px solid #ececec; border-radius: 10px;
        padding: 10px 14px; box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    }
    section[data-testid="stSidebar"] {
        border-right: 1px solid #eee;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# =============================================================================
# PAGE FUNCTIONS
# Each function below is one entry in the sidebar navigation. Streamlit's
# st.navigation/st.Page only executes the SELECTED page's function on each
# rerun — this is the main reason the app should feel noticeably snappier
# than the previous st.tabs version, especially the Charts/Report pages
# which do the most work.
# =============================================================================

def page_how_to_use():
    st.subheader("How to Use This Tool")

    st.markdown("This tool helps you evaluate road pavement condition using two standard indicators — **PCI** (Pavement Condition Index) and **IRI** (International Roughness Index) — and automatically recommends a maintenance action for each road section.")

    st.divider()

    # --- Step 1 ---
    st.markdown("### Step 1 — Prepare your data")
    st.markdown(
        """
Create a spreadsheet or CSV file with **one row per defect** observed on each road section.
Your file must contain these 5 columns:

| Section | Defect Type | Severity | Area Percentage (%) | IRI |
|---|---|---|---|---|
| S1 | Potholes | High | 6 | 3.8 |
| S1 | Raveling | Low | 10 | 3.8 |
| S2 | Longitudinal Crack | Low | 4 | 1.9 |

**Column explanations:**
- **Section** — the road section ID (e.g. S1, S2, S3 ...). If a section has multiple defects, it will appear in multiple rows.
- **Defect Type** — the type of pavement defect observed. Supported types: `Longitudinal Crack`, `Alligator (Fatigue) Crack`, `Potholes`, `Raveling`, `Depression/Sag`, `Patching (Failed)`, `Bleeding/Flushing`, `Rut/Rutting`.
- **Severity** — must be exactly `Low`, `Medium`, or `High`.
- **Area Percentage (%)** — how much of the section area is affected by that defect, as a percentage (e.g. `6` means 6%).
- **IRI (m/km)** — the roughness reading for that section. You can repeat the same IRI value on every defect row for a section, or just fill it in once — the tool will average all IRI values it finds per section.

> **You don't need all columns.** If you only have defect data (no IRI), the tool computes PCI only. If you only have IRI (no defect data), the tool computes IRI only. Both together gives the full combined rating.
        """
    )

    st.divider()

    # --- Step 2 ---
    st.markdown("### Step 2 — Upload your data or enter it manually")
    st.markdown(
        """
You have **two ways** to get data into the tool — choose whichever suits you:

**Option A — Upload a file** using the **Data Input** panel in the left sidebar.
- **Supported formats:** `.csv`, `.xlsx`, `.xls`
- **Also supported:** the lecturer's multi-sheet `TCG633_PCI_IRI_Pro` Excel template — upload it as-is, no reformatting needed. The tool automatically reads the `PCI_Input` and `IRI_Input` sheets.
- **No file yet?** Click **Load Built-in Dataset** in the sidebar to instantly load a pre-filled example dataset (10 road sections, S1–S10) so you can explore all the pages right away.

Once uploaded, the sidebar will show a green confirmation message and the number of rows loaded. Go to **Upload & Preview** in the sidebar to see and confirm your raw data before analysis.
        """
    )

    st.divider()

    # --- Step 2b ---
    st.markdown("### Step 2 (Alternative) — Enter data manually")
    st.markdown(
        """
If you do not have a spreadsheet file ready, you can type your data directly into the tool using the **✏️ Enter Data Manually** page in the sidebar. No file preparation needed.

**How it works:**

1. Open **Enter Data Manually** from the sidebar.
2. Fill in the form at the top of the page:
   - **Section ID** — type the section name, e.g. `S1`, `S2`, `S3`.
   - **Defect Type** — choose from the dropdown (8 defect types supported).
   - **Severity** — choose `Low`, `Medium`, or `High`.
   - **Area Percentage (%)** — how much of the section surface is affected (e.g. `6` means 6%). Set to `0` if you are only recording an IRI reading.
   - **IRI (m/km)** — the roughness value for this section. Set to `0` if you are only recording defects with no roughness reading.
3. Click **➕ Add Row**. The row appears in the table below the form.
4. Repeat for every defect on every section. If a section has more than one defect, add one row per defect — use the same Section ID each time (e.g. add `S1/Potholes/High/6/3.8` as one row, then `S1/Raveling/Low/10/3.8` as a second row).
5. Made a mistake? Use the **Delete selected row** field to remove a specific row by its row number, or **Clear all rows** to start over.
6. When all your data is entered, click **✅ Use This Data for Analysis**. The tool immediately loads your data and it becomes available on the Dashboard, Charts, Hybrid Index, and Report pages.
7. Optionally, click **⬇️ Download as CSV** to save what you entered as a `.csv` file — you can re-upload this file later without needing to re-enter everything.

> **Tip:** you can mix approaches. For example, enter your data manually first to test the tool, then download the CSV, add more rows in Excel, and re-upload it later.
        """
    )

    st.divider()

    # --- Step 3 ---
    st.markdown("### Step 3 — View your results")
    st.markdown(
        """
Once data is loaded, navigate the pages in the sidebar:

| Page | What you get |
|---|---|
| 📊 **Dashboard** | KPI summary cards (total sections, average PCI, average IRI, poor sections count) and a condition distribution chart |
| 📋 **Detailed Results** | Full section-level summary table, defect-level computation breakdown, and download buttons |
| 📈 **Charts** | Interactive bar charts — PCI by section, IRI by section, defect type distribution, and condition rating distribution |
| 🧮 **Hybrid Index** | A single 0-100 score per section that blends PCI and IRI together — see Step 4 below |
| 📄 **Report Generator** | Auto-generates a complete HTML report you can download and print as PDF |
| 📐 **Methodology & Assumptions** | Full explanation of every formula, factor, and assumption used in all calculations |
        """
    )

    st.divider()

    # --- Step 4 ---
    st.markdown("### Step 4 — Explore the Hybrid Index")
    st.markdown(
        """
The **Dashboard**, **Detailed Results**, and **Charts** pages all use the official Combined Condition Rating — a category label (Very Good / Good / Fair / Poor) based on whichever of PCI or IRI is worse for each section.

The **Hybrid Index** page adds a different, optional view: instead of a category label, it produces **one number from 0 to 100 per section**, blending PCI and IRI together into a single score.

**How the Hybrid Index is calculated:**
1. IRI is first converted onto the same 0–100 scale as PCI, using the same classification boundaries (≥85 = Very Good, ≥70 = Good, ≥55 = Fair, below 55 = Poor). This step is needed because IRI normally goes in the opposite direction — a lower IRI is better, while a higher PCI is better.
2. A **weight slider** lets you decide how much PCI matters vs IRI. For example, if you set it to 70%:
   > `Hybrid Index = 0.70 × PCI + 0.30 × IRI Score`
3. The result is classified using the same bands as PCI.

**When to use it:** the Hybrid Index is useful when you want to **rank sections in priority order** (the section with the lowest Hybrid Index needs attention most) or when you want a single number to display in a report chart. It supplements — but does not replace — the official Combined Condition Rating used elsewhere.

**How to use the page:**
1. Open **Hybrid Index** in the sidebar.
2. Move the slider to set how much weight goes to PCI (default is 60%).
3. Read the table and chart. The table also shows the IRI Score (the rescaled 0–100 version of IRI) alongside the original IRI so you can see how the conversion works.
4. Click **Download Hybrid Index (CSV)** to save just this table.
        """
    )

    st.divider()

    # --- Step 5 ---
    st.markdown("### Step 5 — Generate and download a report")
    st.markdown(
        """
The **Report Generator** page builds a complete, self-contained report in one click.

1. Open **Report Generator** in the sidebar.
2. Click **Generate Report**.
3. A preview of the report appears below the button.
4. Click **Download Report (HTML)** to save the file.
5. Open the downloaded `.html` file in any web browser.
6. To get a PDF: in the browser, press **Ctrl+P** (or **Cmd+P** on Mac) → change destination to **Save as PDF** → Save.

The report includes: KPI summary, PCI and IRI charts, full section results table, defect-level detail, Hybrid Index table (if you visited that page first), and a methodology summary — everything you need for a technical report submission.
        """
    )


def page_manual_entry():
    st.subheader("✏️ Enter Data Manually")
    st.markdown(
        "Fill in the form below to add pavement condition data row by row. "
        "Each row represents **one defect** observed in one road section. "
        "When you are done adding all rows, click **✅ Use This Data for Analysis** "
        "to send it to the Dashboard, Charts, and all other pages."
    )

    # Initialise the in-memory row list
    if "manual_rows" not in st.session_state:
        st.session_state["manual_rows"] = []

    # ------------------------------------------------------------------ form
    st.markdown("#### ➕ Add a new row")
    with st.form("manual_entry_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            section_input = st.text_input(
                "Section ID *",
                placeholder="e.g. S1",
                help="The road section identifier. Use the same ID for all defects in the same section.",
            )
        with c2:
            defect_input = st.selectbox(
                "Defect Type *",
                options=[
                    "Longitudinal Crack",
                    "Alligator (Fatigue) Crack",
                    "Potholes",
                    "Raveling",
                    "Depression/Sag",
                    "Patching (Failed)",
                    "Bleeding/Flushing",
                    "Rut/Rutting",
                ],
                help="Select the type of pavement defect observed.",
            )
        with c3:
            severity_input = st.selectbox(
                "Severity *",
                options=["Low", "Medium", "High"],
                help="How severe the defect is.",
            )

        c4, c5 = st.columns(2)
        with c4:
            area_str = st.text_input(
                "Area Percentage (%)",
                value="",
                placeholder="e.g. 6.5",
                help="Percentage of the section area affected by this defect (0–100). Type any number.",
            )
        with c5:
            iri_str = st.text_input(
                "IRI (m/km)",
                value="",
                placeholder="e.g. 3.8",
                help="International Roughness Index reading for this section (m/km). "
                     "Leave blank if you are only recording defect data (no roughness measurement).",
            )

        submitted = st.form_submit_button("➕ Add Row", type="primary", use_container_width=True)

    if submitted:
        # Parse text inputs with friendly errors
        area_input = 0.0
        iri_input = 0.0
        parse_ok = True

        if area_str.strip():
            try:
                area_input = float(area_str.strip())
                if not (0.0 <= area_input <= 100.0):
                    st.error("Area Percentage must be between 0 and 100.")
                    parse_ok = False
            except ValueError:
                st.error(f"Area Percentage — '{area_str}' is not a valid number. Please type a number like 6.5")
                parse_ok = False

        if iri_str.strip():
            try:
                iri_input = float(iri_str.strip())
                if iri_input < 0:
                    st.error("IRI cannot be negative.")
                    parse_ok = False
            except ValueError:
                st.error(f"IRI — '{iri_str}' is not a valid number. Please type a number like 3.8")
                parse_ok = False

        if not parse_ok:
            pass
        elif not section_input.strip():
            st.error("Section ID is required. Please enter a section name (e.g. S1).")
        elif area_input == 0.0 and iri_input == 0.0:
            st.warning(
                "Both Area Percentage and IRI are 0 (or blank). "
                "Please enter at least one non-zero value before adding this row.",
            )
        else:
            new_row = {
                "Section": section_input.strip().upper(),
                "Defect Type": defect_input if area_input > 0 else None,
                "Severity": severity_input if area_input > 0 else None,
                "Area Percentage (%)": area_input if area_input > 0 else None,
                "IRI": iri_input if iri_input > 0 else None,
            }
            st.session_state["manual_rows"].append(new_row)
            st.success(
                f"Row added — Section **{new_row['Section']}**, "
                f"{defect_input} ({severity_input}), "
                f"Area {area_input}%, IRI {iri_input} m/km."
            )

    # ---------------------------------------------------------- current table
    rows = st.session_state["manual_rows"]

    if rows:
        st.markdown(f"#### 📋 Rows entered so far ({len(rows)} total)")

        preview_df = pd.DataFrame(rows)
        preview_df = preview_df.fillna("—")

        # Display with a row-number column so users can identify which to delete
        display_df = preview_df.copy()
        display_df.insert(0, "#", range(1, len(display_df) + 1))
        st.dataframe(display_df, use_container_width=True, hide_index=True, height=min(400, 55 + 35 * len(rows)))

        # Delete a row
        st.markdown("#### 🗑️ Delete a row")
        col_del, col_clr = st.columns([2, 1])
        with col_del:
            row_to_delete = st.number_input(
                "Enter the row number to delete (see # column above)",
                min_value=1, max_value=len(rows), step=1, value=1,
                key="delete_row_num",
            )
            if st.button("🗑️ Delete selected row", use_container_width=True):
                st.session_state["manual_rows"].pop(int(row_to_delete) - 1)
                st.rerun()
        with col_clr:
            st.markdown("&nbsp;")
            if st.button("🧹 Clear all rows", use_container_width=True):
                st.session_state["manual_rows"] = []
                st.session_state.pop("df_raw", None)
                st.session_state.pop("data_source", None)
                st.rerun()

        st.divider()

        # -------------------------------------------------- use / download
        st.markdown("#### ✅ Use this data")
        col_use, col_csv = st.columns(2)
        with col_use:
            if st.button(
                "✅ Use This Data for Analysis",
                type="primary",
                use_container_width=True,
                help="Sends all rows above to the Dashboard, Charts, Hybrid Index, and Report pages.",
            ):
                out_df = pd.DataFrame(rows)
                for col in CANONICAL_COLS:
                    if col not in out_df.columns:
                        out_df[col] = None
                out_df = out_df[CANONICAL_COLS]
                st.session_state["df_raw"] = out_df
                st.session_state["data_source"] = f"Manually entered data ({len(rows)} rows)"
                # Clear any cached computation so it reruns with the new data
                st.session_state.pop("_summary_df", None)
                st.session_state.pop("_detail_df", None)
                st.session_state.pop("_hybrid_df", None)
                st.session_state.pop("_report_html", None)
                st.success(
                    f"✅ Done! {len(rows)} rows are now loaded. "
                    "Go to **Dashboard** or **Charts** in the sidebar to see your results."
                )

        with col_csv:
            out_df_dl = pd.DataFrame(rows)
            for col in CANONICAL_COLS:
                if col not in out_df_dl.columns:
                    out_df_dl[col] = None
            out_df_dl = out_df_dl[CANONICAL_COLS]
            st.download_button(
                "⬇️ Download as CSV",
                data=out_df_dl.to_csv(index=False).encode("utf-8"),
                file_name=f"manual_entry_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True,
                help="Download your entered rows as a CSV file that you can re-upload later.",
            )

    else:
        st.info(
            "No rows added yet. Fill in the form above and click **➕ Add Row** "
            "to start building your dataset.",
            icon="👆",
        )

    # ---------------------------------------------------------- quick guide
    with st.expander("📖 Tips for filling in the form", expanded=False):
        st.markdown(
            """
**Section ID** — Use a short label like `S1`, `S2`, `S3` etc. All defects found on the
same stretch of road should share the same Section ID. There is no limit on how many rows
you can add per section.

**Defect Type** — Choose from the dropdown. These match the 8 types recognised by the
PCI calculation model. If your defect isn't in the list, choose the closest one and note
it in your report.

**Severity** — Choose `Low`, `Medium`, or `High` based on how badly the defect is
affecting the road surface.

**Area Percentage (%)** — Estimate what fraction of the section's surface area shows
this defect. For example, if roughly 1/10 of the road is cracked, enter `10`.
Set to `0` if you are only recording an IRI reading for this row (no defect).

**IRI (m/km)** — The roughness value measured by a profilometer or similar device for
this section. If you are only recording defects (no roughness measurement available),
leave this as `0`. You can also repeat the same IRI value on every row of a section —
the tool will average all values per section automatically.

**Adding multiple defects per section:** simply add one row per defect, using the same
Section ID. Example — S1 has two defects: add S1/Potholes/High/6/3.8 as one row, then
add S1/Raveling/Low/10/3.8 as a second row.

**Saving your work:** click **⬇️ Download as CSV** at any time to save your entered rows
as a file. You can re-upload this CSV later from the sidebar without needing to re-enter
everything.
            """
        )


def page_upload_preview():
    st.subheader("Upload & Preview")
    df_raw = st.session_state.get("df_raw")
    if df_raw is None:
        st.warning("No data loaded yet. Use the sidebar to upload a file or load the built-in dataset.")
        return
    missing = [c for c in CANONICAL_COLS if c not in df_raw.columns]
    if "Section" in missing:
        st.error(
            "Your file must contain a 'Section' column. Detected columns: "
            + ", ".join(map(str, df_raw.columns))
        )
        return
    if missing:
        st.warning(
            f"Columns not found and treated as empty: {', '.join(missing)}. "
            "The tool will still compute whatever indicator(s) the available data supports."
        )
    st.success(f"Data validated — {df_raw['Section'].nunique()} unique section(s), {len(df_raw)} row(s).")
    st.dataframe(df_raw, use_container_width=True, height=420)


def page_dashboard():
    st.subheader("Summary Dashboard")
    summary_df = st.session_state.get("_summary_df")
    calc_flags = st.session_state.get("_calc_flags", [])
    if summary_df is None or summary_df.empty:
        st.info("Load data to see the dashboard.")
        return

    if calc_flags:
        with st.expander(f"⚠️ {len(calc_flags)} data-quality flag(s) detected — click to view", expanded=False):
            for f in sorted(set(calc_flags)):
                st.write("• " + f)

    total_sections = len(summary_df)
    avg_pci = summary_df["PCI"].mean()
    avg_iri = summary_df["Avg IRI (m/km)"].mean()
    poor_sections = int((summary_df["Combined Condition Rating"].isin(["Poor", "Poor (Rough)"])).sum())
    valid_ranks = summary_df["_rank"].dropna()
    overall_rank = int(round(valid_ranks.mean())) if not valid_ranks.empty else None
    overall_label = PCI_RANK_LABEL.get(overall_rank, "No Data") if overall_rank else "No Data"

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        st.metric("Total Road Sections", total_sections)
    with c2:
        st.metric("Average PCI", f"{avg_pci:.1f}" if pd.notna(avg_pci) else "N/A")
    with c3:
        st.metric("Average IRI (m/km)", f"{avg_iri:.2f}" if pd.notna(avg_iri) else "N/A")
    with c4:
        st.metric("Poor Sections", poor_sections)
    with c5:
        st.markdown("**Overall Network Condition**")
        st.markdown(condition_badge(overall_label), unsafe_allow_html=True)

    st.markdown("")
    colL, colR = st.columns([0.55, 0.45])
    with colL:
        st.markdown("##### Section Condition Overview")
        show_cols = ["Section", "PCI", "PCI Condition", "Avg IRI (m/km)", "IRI Condition",
                     "Combined Condition Rating", "Maintenance Recommendation"]
        st.dataframe(style_dataframe(summary_df[show_cols]), use_container_width=True, height=380)
    with colR:
        st.markdown("##### Condition Rating Distribution")
        dist = summary_df["Combined Condition Rating"].value_counts().reset_index()
        dist.columns = ["Condition", "Count"]
        if PLOTLY_OK:
            fig = px.pie(
                dist, names="Condition", values="Count", hole=0.45,
                color="Condition", color_discrete_map=CONDITION_COLORS,
            )
            fig.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=360)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.bar_chart(dist.set_index("Condition")["Count"])


def page_detailed_results():
    st.subheader("Detailed Results")
    summary_df = st.session_state.get("_summary_df")
    detail_df = st.session_state.get("_detail_df")
    df_raw = st.session_state.get("df_raw")
    if summary_df is None or summary_df.empty:
        st.info("Load data to see detailed results.")
        return

    # ── Section filter ────────────────────────────────────────────────────────
    all_sections = list(summary_df["Section"])
    selected_sections = st.multiselect(
        "🔍 Filter by Section — select one or more (leave blank to show all)",
        options=all_sections,
        default=[],
        placeholder="All sections shown — click here to filter",
        key="detail_section_filter",
    )
    active_sections = selected_sections if selected_sections else all_sections

    display_summary = summary_df.drop(columns=["_rank"])
    filtered_summary = display_summary[display_summary["Section"].isin(active_sections)]

    if selected_sections:
        st.caption(f"Showing {len(filtered_summary)} of {len(display_summary)} sections.")

    st.markdown("##### 1. Section-Level Summary (PCI, IRI, Combined Rating, Recommendation)")
    st.dataframe(style_dataframe(filtered_summary), use_container_width=True,
                 height=min(600, 80 + 35 * len(filtered_summary)))

    st.markdown("##### 2. Defect-Level Detail & Computation")
    if detail_df is None or detail_df.empty:
        st.info("No defect-level rows found in the uploaded data (Defect Type / Severity / Area % missing).")
    else:
        detail_show = detail_df[[
            "Section", "Defect Type", "Severity", "Area Percentage (%)",
            "Weighting Factor", "Severity Factor", "Deduct Value", "Suggested Defect Treatment"
        ]].reset_index(drop=True)
        filtered_detail = detail_show[detail_show["Section"].isin(active_sections)]
        st.dataframe(style_dataframe(filtered_detail), use_container_width=True,
                     height=min(500, 80 + 35 * len(filtered_detail)))
        st.caption(
            "Deduct Value = Area (%) × Severity Factor × Weighting Factor. "
            "'Suggested Defect Treatment' is supplementary general guidance per "
            "defect (see Methodology page) — the official section rating/recommendation "
            "is the 'Combined Condition Rating' / 'Maintenance Recommendation' columns above."
        )

    st.markdown("##### 3. Download Analysed Results")
    c1, c2 = st.columns(2)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    with c1:
        st.download_button(
            "⬇️ Download Section Summary (CSV)",
            data=display_summary.to_csv(index=False).encode("utf-8"),
            file_name=f"pavement_section_summary_{ts}.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with c2:
        if not OPENPYXL_OK:
            st.button(
                "⬇️ Download Full Results (Excel) — unavailable",
                disabled=True, use_container_width=True,
                help="The 'openpyxl' package isn't available in this server environment right now.",
            )
            st.caption(
                "⚠️ Excel download is temporarily unavailable on this server "
                "(missing 'openpyxl' package) — use the CSV download on the left instead."
            )
        else:
            try:
                excel_bytes = df_to_excel_bytes({
                    "Section_Summary": display_summary,
                    "Defect_Detail": detail_df if detail_df is not None else pd.DataFrame(),
                    "Raw_Input": df_raw,
                })
                st.download_button(
                    "⬇️ Download Full Results (Excel, 3 sheets)",
                    data=excel_bytes,
                    file_name=f"pavement_analysis_results_{ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.caption(f"⚠️ Excel export failed ({e}). Use the CSV download on the left instead.")

    if not OPENPYXL_OK:
        with st.expander("ℹ️ Why is Excel upload/download unavailable? (click to view)"):
            st.markdown(
                "This server's Python environment is missing the **openpyxl** package, "
                "which both Excel *upload* and Excel *download* depend on. This is almost "
                "always a deployment/server-side issue, not a problem with your data.\n\n"
                "**If you're the app owner (Streamlit Community Cloud):** delete the app and "
                "redeploy it, choosing Python 3.11 or 3.12 in **Advanced settings** before "
                "clicking Deploy. In the meantime, CSV upload/download both work normally."
            )


def page_charts():
    st.subheader("Charts")
    summary_df = st.session_state.get("_summary_df")
    detail_df = st.session_state.get("_detail_df")
    if summary_df is None or summary_df.empty:
        st.info("Load data to see charts.")
        return

    if not PLOTLY_OK:
        st.warning(
            "Plotly isn't installed in this environment, so charts below use "
            "Streamlit's simplified built-in charts. Add `plotly` to "
            "requirements.txt and reboot the app for the full interactive charts.",
            icon="⚠️",
        )

    st.markdown("##### PCI by Section")
    pci_chart_df = summary_df.dropna(subset=["PCI"])
    if pci_chart_df.empty:
        st.caption("No PCI data available to chart.")
    elif PLOTLY_OK:
        fig1 = px.bar(
            pci_chart_df, x="Section", y="PCI", color="PCI Condition",
            color_discrete_map=CONDITION_COLORS, text="PCI",
            category_orders={"Section": list(summary_df["Section"])},
        )
        fig1.add_hline(y=85, line_dash="dot", line_color="#2E7D32", annotation_text="Very Good ≥85")
        fig1.add_hline(y=70, line_dash="dot", line_color="#1976D2", annotation_text="Good ≥70")
        fig1.add_hline(y=55, line_dash="dot", line_color="#F9A825", annotation_text="Fair ≥55")
        fig1.update_traces(texttemplate="%{text:.1f}", textposition="outside")
        fig1.update_layout(yaxis_range=[0, 105], height=380, margin=dict(t=30, b=10))
        st.plotly_chart(fig1, use_container_width=True)
    else:
        st.bar_chart(pci_chart_df.set_index("Section")["PCI"])

    st.markdown("##### IRI by Section")
    iri_chart_df = summary_df.dropna(subset=["Avg IRI (m/km)"])
    if iri_chart_df.empty:
        st.caption("No IRI data available to chart.")
    elif PLOTLY_OK:
        fig2 = px.bar(
            iri_chart_df, x="Section", y="Avg IRI (m/km)", color="IRI Condition",
            color_discrete_map=CONDITION_COLORS, text="Avg IRI (m/km)",
            category_orders={"Section": list(summary_df["Section"])},
        )
        fig2.add_hline(y=2, line_dash="dot", line_color="#2E7D32", annotation_text="Very Good <2")
        fig2.add_hline(y=3, line_dash="dot", line_color="#1976D2", annotation_text="Good <3")
        fig2.add_hline(y=4, line_dash="dot", line_color="#F9A825", annotation_text="Fair <4")
        fig2.update_traces(texttemplate="%{text:.2f}", textposition="outside")
        fig2.update_layout(height=380, margin=dict(t=30, b=10))
        st.plotly_chart(fig2, use_container_width=True)
    else:
        st.bar_chart(iri_chart_df.set_index("Section")["Avg IRI (m/km)"])

    colA, colB = st.columns(2)
    with colA:
        st.markdown("##### Defect Type Distribution")
        if detail_df is None or detail_df.empty:
            st.caption("No defect-level data available to chart.")
        else:
            defect_counts = detail_df.groupby(["Defect Type", "Severity"]).size().reset_index(name="Count")
            if PLOTLY_OK:
                fig3 = px.bar(
                    defect_counts, x="Defect Type", y="Count", color="Severity",
                    color_discrete_map={"Low": "#2E7D32", "Medium": "#F9A825", "High": "#C62828"},
                    barmode="stack",
                )
                fig3.update_layout(height=380, margin=dict(t=10, b=10), xaxis_tickangle=-30)
                st.plotly_chart(fig3, use_container_width=True)
            else:
                pivot = defect_counts.pivot_table(index="Defect Type", columns="Severity", values="Count", fill_value=0)
                st.bar_chart(pivot)

    with colB:
        st.markdown("##### Condition Rating Distribution")
        dist2 = summary_df["Combined Condition Rating"].value_counts().reset_index()
        dist2.columns = ["Condition", "Count"]
        if PLOTLY_OK:
            fig4 = px.bar(
                dist2, x="Condition", y="Count", color="Condition",
                color_discrete_map=CONDITION_COLORS, text="Count",
            )
            fig4.update_layout(height=380, margin=dict(t=10, b=10), showlegend=False)
            st.plotly_chart(fig4, use_container_width=True)
        else:
            st.bar_chart(dist2.set_index("Condition")["Count"])

    st.markdown("##### Defects by Section (stacked)")
    if detail_df is None or detail_df.empty:
        st.caption("No defect-level data available to chart.")
    elif PLOTLY_OK:
        by_section = detail_df.groupby(["Section", "Defect Type"]).size().reset_index(name="Count")
        fig5 = px.bar(
            by_section, x="Section", y="Count", color="Defect Type", barmode="stack",
            category_orders={"Section": list(summary_df["Section"])},
        )
        fig5.update_layout(height=380, margin=dict(t=10, b=10))
        st.plotly_chart(fig5, use_container_width=True)
    else:
        by_section = detail_df.groupby(["Section", "Defect Type"]).size().reset_index(name="Count")
        pivot2 = by_section.pivot_table(index="Section", columns="Defect Type", values="Count", fill_value=0)
        st.bar_chart(pivot2)


def page_hybrid_index():
    st.subheader("🧮 Hybrid Index (Bonus)")
    summary_df = st.session_state.get("_summary_df")
    if summary_df is None or summary_df.empty:
        st.info("Load data to compute a hybrid index.")
        return

    st.markdown(
        "This combines PCI and IRI into a **single 0-100 number per section**, "
        "in addition to (not replacing) the official Combined Condition Rating "
        "used elsewhere in the app. IRI is rescaled onto the same 0-100 scale "
        "as PCI using the identical classification boundaries (85/70/55), then "
        "blended with PCI using the weight you choose below."
    )

    w_pci_pct = st.slider(
        "Weight given to PCI (remaining weight goes to IRI)",
        min_value=0, max_value=100, value=60, step=5, format="%d%%",
        key="hybrid_w_pci",
    )
    w_pci = w_pci_pct / 100.0
    st.caption(f"Current blend: **{w_pci_pct}% PCI + {100 - w_pci_pct}% IRI**")

    hybrid_df = compute_hybrid_index(summary_df, w_pci)
    st.session_state["_hybrid_df"] = hybrid_df  # cache for Report page

    show_cols = ["Section", "PCI", "Avg IRI (m/km)", "IRI Score (0-100)", "Hybrid Index", "Hybrid Condition"]
    st.dataframe(style_dataframe(hybrid_df[show_cols]), use_container_width=True, height=340)

    st.markdown("##### Hybrid Index by Section")
    chart_df = hybrid_df.dropna(subset=["Hybrid Index"])
    if chart_df.empty:
        st.caption("No data to chart.")
    elif PLOTLY_OK:
        fig = px.bar(
            chart_df, x="Section", y="Hybrid Index", color="Hybrid Condition",
            color_discrete_map=CONDITION_COLORS, text="Hybrid Index",
            category_orders={"Section": list(summary_df["Section"])},
        )
        fig.update_traces(texttemplate="%{text:.1f}", textposition="outside")
        fig.update_layout(yaxis_range=[0, 105], height=380, margin=dict(t=20, b=10))
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.bar_chart(chart_df.set_index("Section")["Hybrid Index"])

    csv_bytes = hybrid_df[show_cols].to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Download Hybrid Index (CSV)",
        data=csv_bytes,
        file_name=f"hybrid_index_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
        mime="text/csv",
    )


def page_report_generator():
    st.subheader("📄 Automated Report Generator (Bonus)")
    summary_df = st.session_state.get("_summary_df")
    detail_df = st.session_state.get("_detail_df")
    calc_flags = st.session_state.get("_calc_flags", [])
    data_source = st.session_state.get("data_source")
    hybrid_df = st.session_state.get("_hybrid_df")

    if summary_df is None or summary_df.empty:
        st.info("Load data first to generate a report.")
        return

    st.markdown(
        "Generates a single self-contained **HTML report** — KPIs, charts, "
        "section summary, defect detail, and (if you've visited the Hybrid "
        "Index page) the hybrid index table — ready to download and attach "
        "to your technical report submission."
    )
    st.caption(
        "💡 No PDF library is used here on purpose (keeps the app dependency-free "
        "and crash-resistant). Open the downloaded .html file in any browser, "
        "then use **File → Print → Save as PDF** to get a PDF version in seconds."
    )

    if st.button("🛠️ Generate Report", type="primary"):
        with st.spinner("Building report..."):
            html = build_html_report(summary_df, detail_df, data_source, calc_flags, hybrid_df)
        st.session_state["_report_html"] = html
        st.success("Report generated below — preview it, then download.")

    html = st.session_state.get("_report_html")
    if html:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        st.download_button(
            "⬇️ Download Report (HTML)",
            data=html.encode("utf-8"),
            file_name=f"pavement_condition_report_{ts}.html",
            mime="text/html",
            use_container_width=True,
        )
        with st.expander("👁️ Preview report", expanded=True):
            st.components.v1.html(html, height=600, scrolling=True)


def page_methodology():
    st.subheader("Methodology & Assumptions")
    st.markdown(
        """
This tool follows the computation logic encoded in the lecturer-provided
workbooks `TCG633_PCI_IRI_Model.xlsx` and `TCG633_PCI_IRI_Pro_v2.xlsx`
(see their `Lookup`, `PCI_Compute`, `IRI_Compute`, and `Settings_Summary`
sheets). It is a **teaching simplification** of pavement evaluation practice,
not the full ASTM D6433 deduct-curve procedure.

### 1. PCI (Pavement Condition Index) — simplified deduct-value method

```
Deduct Value = Area Affected (%) × Severity Factor × Weighting Factor
```

| Defect Type | Weighting Factor | | Severity | Factor |
|---|---|---|---|---|
| Longitudinal Crack | 1.0 | | Low | 0.6 |
| Alligator (Fatigue) Crack | 1.6 | | Medium | 1.0 |
| Potholes | 2.2 | | High | 1.4 |
| Raveling | 1.2 | | | |
| Depression/Sag | 1.4 | | | |
| Patching (Failed) | 1.8 | | | |
| Bleeding/Flushing | 1.0 | | | |
| Rut/Rutting | 1.6 | | | |

For each **Section**, all Deduct Values are summed, then:

```
PCI = MAX( 0, 100 − MIN(100, ΣDeduct Value) )
```

**PCI condition classes:**

| PCI Range | Condition | Recommendation |
|---|---|---|
| 85–100 | Very Good | Routine maintenance |
| 70–84 | Good / Satisfactory | Preventive maintenance (crack sealing, local patching) |
| 55–69 | Fair | Surface treatment / overlay (localized) |
| 0–54 | Poor | Major rehabilitation / reconstruction assessment |

### 2. IRI (International Roughness Index)

```
Average IRI = mean(IRI readings for that section)
```

**IRI condition classes:**

| IRI (m/km) | Condition | Recommendation |
|---|---|---|
| < 2 | Very Good (Smooth) | Routine maintenance |
| 2 – < 3 | Good | Preventive maintenance (localized patching/leveling) |
| 3 – < 4 | Fair | Surface treatment / thin overlay |
| ≥ 4 | Poor (Rough) | Structural overlay / rehabilitation |

### 3. Combined (Hybrid) Condition Rating — official, used everywhere by default

Following the "Hybrid mode" logic in `TCG633_PCI_IRI_Pro_v2.xlsx`
(`Settings_Summary` sheet): when **both** PCI and IRI are available for a
section, the **worse (more severe)** of the two condition classes governs
the combined rating and its recommendation. If only one indicator is
available for a section, that indicator alone determines the rating.

### 4. Hybrid Index (bonus page) — numeric blend, additional to #3

The **Hybrid Index** page adds a second, optional way of combining PCI and
IRI: instead of taking the worse classification, it rescales IRI onto the
same 0-100 scale as PCI (using the identical 85/70/55 boundary points), then
computes a weighted average:

```
IRI Score = piecewise-linear rescale of IRI onto 0-100,
            anchored at (0→100, 2→85, 3→70, 4→55, 6→0)
Hybrid Index = w × PCI + (1 − w) × IRI Score      (w adjustable in-app, default 60% PCI)
```

This is clearly separated from the official rating in #3 throughout the app
and in the downloadable report.

### 5. Automated Report Generation (bonus page)

Builds a single self-contained HTML file (KPIs, inline SVG charts, full
tables) with no additional PDF library — keeping the app's dependency
footprint minimal and resistant to the kind of deployment breakage seen with
optional packages on some hosting platforms. Open the file in any browser
and use Print → Save as PDF for a PDF copy.

### 6. Other assumptions made explicit by this tool

- **Unrecognised defect type** → neutral weighting factor of **1.0** substituted, flagged on Dashboard.
- **Unrecognised severity** → treated as **Medium (factor 1.0)**, flagged.
- **Section with defect data but no IRI** → rated on **PCI only**.
- **Section with IRI but no defect rows** → rated on **IRI only**.
- **Per-defect "Suggested Defect Treatment"** column is supplementary general
  guidance added by this tool, not part of the lecturer's spreadsheet, and is
  kept visually separate from the official section-level recommendation.
- Condition-class thresholds and recommendation wording are reproduced
  exactly as given in the Lookup sheet of the supplied workbooks.
        """
    )
    st.warning(
        "This is a teaching/learning tool built for an academic assignment. "
        "It should not be used for real-world pavement asset management "
        "decisions without validation against the full ASTM D6433 / JKR "
        "manual procedures by a qualified engineer.",
        icon="⚠️",
    )


# =============================================================================
# SIDEBAR — NAVIGATION + DATA INPUT
# =============================================================================
with st.sidebar:
    pg = st.navigation(
        [
            st.Page(page_how_to_use, title="How to Use", icon="🏠", default=True),
            st.Page(page_upload_preview, title="Upload & Preview", icon="📥"),
            st.Page(page_manual_entry, title="Enter Data Manually", icon="✏️"),
            st.Page(page_dashboard, title="Dashboard", icon="📊"),
            st.Page(page_detailed_results, title="Detailed Results", icon="📋"),
            st.Page(page_charts, title="Charts", icon="📈"),
            st.Page(page_hybrid_index, title="Hybrid Index", icon="🧮"),
            st.Page(page_report_generator, title="Report Generator", icon="📄"),
            st.Page(page_methodology, title="Methodology & Assumptions", icon="📐"),
        ]
    )

    st.divider()
    st.header("📥 Data Input")
    uploaded_file = st.file_uploader(
        "Upload pavement condition data (.csv, .xlsx, .xls)",
        type=["csv", "xlsx", "xls"],
        help="Required columns: Section, Defect Type, Severity, Area Percentage (%), IRI",
    )

    col_a, col_b = st.columns(2)
    with col_a:
        load_sample = st.button("📊 Load Built-in Dataset", use_container_width=True)
    with col_b:
        clear_data = st.button("🗑️ Clear", use_container_width=True)

    if clear_data:
        st.session_state.pop("df_raw", None)
        st.session_state.pop("data_source", None)
        st.session_state.pop("_hybrid_df", None)
        st.session_state.pop("_report_html", None)

    if uploaded_file is not None:
        try:
            st.session_state["df_raw"] = read_uploaded_file(uploaded_file)
            st.session_state["data_source"] = f"Uploaded file: {uploaded_file.name}"
        except Exception as e:
            st.error(
                f"File received, but it could not be parsed: {e}\n\n"
                "(The file chip above only confirms it was transferred to the app — "
                "this error means the server-side step that opens/reads it failed.)"
            )

    if load_sample:
        st.session_state["df_raw"] = generate_sample_data()
        st.session_state["data_source"] = "Built-in dataset — TCG633 Sections 1–10 (from your completed workbook)"

    st.divider()
    if "df_raw" in st.session_state:
        st.success(st.session_state.get("data_source", "Data loaded"))
        st.caption(f"{len(st.session_state['df_raw'])} rows loaded")
    else:
        st.info("Upload a file or load the built-in dataset to begin.")

    st.divider()
    st.caption(
        "Required columns:\n"
        "- Section\n- Defect Type\n- Severity (Low/Medium/High)\n"
        "- Area Percentage (%)\n- IRI (m/km)\n\n"
        "A section may appear in several rows (one per defect). "
        "IRI may be repeated per row or given once per section."
    )

# -----------------------------------------------------------------------------
# RUN ANALYSIS ONCE PER RERUN (shared across whichever page is selected)
# -----------------------------------------------------------------------------
df_raw = st.session_state.get("df_raw")
detail_df, summary_df, calc_flags = (None, None, [])
if df_raw is not None and "Section" in df_raw.columns:
    detail_df, summary_df, calc_flags = compute_results(df_raw)
st.session_state["_detail_df"] = detail_df
st.session_state["_summary_df"] = summary_df
st.session_state["_calc_flags"] = calc_flags

# -----------------------------------------------------------------------------
# HEADER  (shown above every page)
# -----------------------------------------------------------------------------
st.title("🛣️ Digital Pavement Condition Evaluation and Maintenance Decision Tool")
st.caption(
    "TCG633 Bridge & Road Maintenance · Individual Project · "
    "Fakulti Kejuruteraan Awam, UiTM Cawangan Sarawak"
)
st.divider()

# -----------------------------------------------------------------------------
# RUN THE SELECTED PAGE
# -----------------------------------------------------------------------------
pg.run()

# -----------------------------------------------------------------------------
# FOOTER  (shown below every page)
# -----------------------------------------------------------------------------
st.divider()
st.caption(
    "TCG633 Bridge & Road Maintenance · Digital Pavement Condition Evaluation and "
    "Maintenance Decision Tool · Built with Streamlit · "
    "Computation logic sourced from lecturer-provided Excel model."
)
